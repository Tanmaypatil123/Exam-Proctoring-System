from django.shortcuts import render

import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from django.http import JsonResponse
from .models import StudentData, DemoFile


def upload_students_details(request):
    demofile = DemoFile.objects.filter(is_active=True).first()
    if request.method == "POST":
        file = request.FILES["file"]

        if file.name.endswith(".csv"):
            # Handle CSV file
            df = pd.read_csv(file)

        elif file.name.endswith(".xlsx"):
            # Handle Excel file
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
            ws = wb.active
            data = []

            for row in ws.iter_rows(values_only=True):
                data.append(row)

            if len(data) < 2:
                return JsonResponse({"message": "Excel file is empty"})

            header = data[0]
            data = data[1:]

            df = pd.DataFrame(data, columns=header)

        else:
            return render(
                request,
                "upload_students_data.html",
                {"message": "Unsupported file format"},
            )

        # Extract specific columns
        if all(
            col in df.columns
            for col in ["First Name", "Last Name", "Email", "Passing Year"]
        ):
            selected_data = df[["First Name", "Last Name", "Email", "Passing Year"]]

            # Convert DataFrame rows to StudentData model instances and save them
            for index, row in selected_data.iterrows():
                pass
                student_data = StudentData(
                    first_name=row["First Name"],
                    last_name=row["Last Name"],
                    email=row["Email"],
                    passing_year=row["Passing Year"],
                )
                student_data.save()
                # proccess to send the message

            return render(
                request,
                "upload_students_data.html",
                {"message": "File uploaded and data saved successfully"},
            )

        else:
            return render(
                request,
                "upload_students_data.html",
                {"message": "Required columns not found in the file"},
            )

    return render(request, "upload_students_data.html", {"demofile": demofile})
